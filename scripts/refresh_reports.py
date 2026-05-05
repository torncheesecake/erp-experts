#!/usr/bin/env python3
"""Refresh marketing reports data from GA4 and the latest LinkedIn export."""

from __future__ import annotations

import csv
import argparse
import json
import os
import re
import shutil
import sys
import time
import urllib.parse
from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime, timedelta
from pathlib import Path
from urllib.parse import urlparse

import openpyxl
import requests
from google.auth.transport.requests import Request
from google.oauth2 import service_account


REPO_ROOT = Path(__file__).resolve().parent.parent
REPORTS_PATH = REPO_ROOT / "src" / "data" / "reports.json"
EXPORTS_DIR = REPO_ROOT / "src" / "exports"
DOWNLOADS_DIR = Path.home() / "Downloads"
PLAYWRIGHT_DOWNLOADS_DIR = REPO_ROOT / ".playwright-cli"
DEFAULT_SERVICE_ACCOUNT_PATH = REPO_ROOT / "credentials" / "google-service-account.json"
SERVICE_ACCOUNT_PATH = Path(
    os.getenv(
        "GOOGLE_SERVICE_ACCOUNT_JSON",
        os.getenv(
            "GOOGLE_APPLICATION_CREDENTIALS",
            str(DEFAULT_SERVICE_ACCOUNT_PATH),
        ),
    )
).expanduser()
GA4_PROPERTY_ID = os.getenv("GA4_PROPERTY_ID", "309916265")
SEARCH_CONSOLE_SITE = "https://www.erpexperts.co.uk/"
EMAIL_WAIT_MINUTES = max(0, int(os.getenv("EMAIL_WAIT_MINUTES", "0")))
EMAIL_POLL_SECONDS = max(5, int(os.getenv("EMAIL_POLL_SECONDS", "20")))
EMAIL_INTERACTIVE_WAIT = os.getenv("EMAIL_INTERACTIVE_WAIT", "0").strip().lower() in {
    "1",
    "true",
    "yes",
}
EMAIL_FILE_EXTENSIONS = {".csv", ".json", ".xlsx"}
EMAIL_FILENAME_HINTS = (
    "email",
    "mailchimp",
    "campaign",
    "newsletter",
    "sendgrid",
    "mailerlite",
    "brevo",
    "tim",
)
LINKEDIN_EXPORT_RE = re.compile(
    r"^Content_(\d{4}-\d{2}-\d{2})_(\d{4}-\d{2}-\d{2})_RicWilson(?: \(\d+\))?\.xlsx$"
)
LINKEDIN_AUDIENCE_EXPORT_RE = re.compile(
    r"^Audience_(\d{4}-\d{2}-\d{2})_(\d{4}-\d{2}-\d{2})_RicWilson(?: \(\d+\))?\.xlsx$"
)
EMAIL_FIELD_ALIASES = {
    "campaignsSent": [
        "campaignssent",
        "campaigns",
        "campaigncount",
        "campaignssentcount",
        "emailssent",
        "totalsent",
        "sends",
    ],
    "totalRecipients": [
        "totalrecipients",
        "recipients",
        "recipientcount",
        "deliveries",
        "delivered",
        "totalaudience",
        "totalsubscribers",
        "totalsent",
    ],
    "openRate": [
        "openrate",
        "uniqueopenrate",
        "openpercentage",
        "openpct",
        "open",
        "opensrate",
    ],
    "clickRate": [
        "clickrate",
        "uniqueclickrate",
        "clickthroughrate",
        "ctr",
        "clickpercentage",
        "clickpct",
    ],
    "unsubscribes": [
        "unsubscribes",
        "unsubscribe",
        "unsubscribed",
        "unsubs",
        "optouts",
        "optout",
    ],
}

INTERNAL_PATH_PREFIXES = ("/reports", "/changelog", "/seo-roadmap", "/api")
TITLE_ALIASES = {
    "ERP Experts | UK's Trusted NetSuite Implementation Partner": "Homepage",
    "About Us | ERP Experts": "About Us",
    "Resources | ERP Experts": "Resources",
    "Contact Us | ERP Experts": "Contact",
    "Case Studies | ERP Experts": "Case Studies",
    "NetSuite Implementation | ERP Experts": "Implementation",
    "Support Plans | ERP Experts": "Support",
    "One Score to Rule Them All": "One Score to Rule Them All",
    "Partners | ERP Experts": "Partners",
    "What is an ERP System? A Beginner's Guide | ERP Experts": "What is an ERP System",
    "What is NetSuite? | Cloud ERP for UK Businesses | ERP Experts": "What is NetSuite?",
}
CTA_LABEL_ALIASES = {
    "navbar_lets_talk": "Navbar - Let's Talk",
    "navbar_mobile_lets_talk": "Navbar Mobile - Let's Talk",
    "hero_start_project": "Hero - Start a Project",
    "footer_cta_book_call": "Footer CTA - Book a Call",
    "footer_cta_start_project": "Footer CTA - Start a Project",
    "footer_start_conversation": "Footer - Start a Conversation",
    "contact_form_submit": "Contact - Form Submit",
    "contact_book_call": "Contact - Book a Call",
    "contact_netscore_cta": "Contact - NETscore CTA",
    "implementation_footer_book_call": "Implementation - Book a Call",
    "implementation_netscore_cta": "Implementation - NETscore CTA",
    "support_footer_book_call": "Support - Book a Call",
    "support_footer_get_support": "Support - Get Support",
    "support_netscore_cta": "Support - NETscore CTA",
    "journey_get_started": "Journey - Get Started",
    "journey_explore_support": "Journey - Explore Support",
    "about_footer_start_conversation": "About - Start a Conversation",
    "about_netscore_cta": "About - NETscore CTA",
    "resources_cta_contact": "Resources - Contact",
    "resources_cta_case_studies": "Resources - Case Studies",
    "resource_article_contact": "Resource Article - Contact",
    "faq_cta_contact": "FAQ - Contact",
    "faq_cta_book_call": "FAQ - Book a Call",
    "faq_netscore_cta": "FAQ - NETscore CTA",
    "project_delivery_hero_cta": "Project Delivery - Hero CTA",
    "project_delivery_hero_audit_cta": "Project Delivery - Hero Audit CTA",
    "project_delivery_footer_contact": "Project Delivery - Footer Contact",
    "project_delivery_footer_audit": "Project Delivery - Footer Audit",
    "project_delivery_netscore_cta": "Project Delivery - NETscore CTA",
    "project_delivery_guarantee_cta": "Project Delivery - Guarantee CTA",
    "implementation_hero_whats_netsuite": "Implementation - What is NetSuite?",
    "partners_cta_case_studies": "Partners - Case Studies",
}
GSC_FALLBACK_PAGES_CSV = REPO_ROOT / "src" / "exports" / "gsc-export" / "Pages.csv"
GA_EVENTS_FALLBACK_CSV = REPO_ROOT / "src" / "exports" / "Events_Event_name.csv"
GA_CHANNELS_FALLBACK_CSV = (
    REPO_ROOT
    / "src"
    / "exports"
    / "Traffic_acquisition_Session_primary_channel_group_(Default_channel_group).csv"
)
GA_PAGES_FALLBACK_CSV = (
    REPO_ROOT / "src" / "exports" / "Pages_and_screens_Page_path_and_screen_class.csv"
)
GA_COUNTRIES_FALLBACK_CSV = REPO_ROOT / "src" / "exports" / "Demographic_details_Country.csv"
ROADMAP_PHASES_TEMPLATE = [
    {
        "id": 1,
        "label": "Phase 1",
        "title": "Quick Wins",
        "subtitle": "Already ranking well — rewrites can drive clicks within weeks",
        "items": [
            {
                "priority": "1a",
                "title": "6 Essential Accounts Receivable Reports",
                "status": "done",
                "why": "Closest to page 1, already getting clicks. A strong rewrite could break top 10 fast. Niche topic = less competition.",
                "keywords": [
                    "accounts receivable reports",
                    "accounts receivable reporting",
                    "AR reports",
                    "accounts receivable report example",
                ],
                "sourcePaths": [
                    "/post/6-essential-accounts-receivable-reports-for-business",
                    "/post/global-payex-app-brings-algoriq-machine-learning-and-ai-to-your-netsuite-accounts-receivables",
                ],
            },
            {
                "priority": "1b",
                "title": "CSV Imports: Why They Fail and How to Fix It",
                "status": "done",
                "why": "Already on page 1. Technical credibility piece that shows E3 knows NetSuite inside out.",
                "keywords": [
                    "netsuite csv import errors",
                    "csv import failure",
                    "netsuite data import",
                ],
                "sourcePaths": [
                    "/post/why-every-30-minute-csv-import-is-one-special-character-away-from-failure-and-how-to-fix-it",
                ],
            },
            {
                "priority": "1c",
                "title": "10 Tips to Maximise Profits",
                "status": "done",
                "why": "Just off page 1. A light refresh can still push it over.",
                "keywords": [
                    "maximise profits",
                    "how to increase profits",
                    "maximising profits",
                ],
                "sourcePaths": ["/post/10-tips-to-maximise-profits"],
            },
        ],
    },
    {
        "id": 2,
        "label": "Phase 2",
        "title": "High-Volume Content",
        "subtitle": "Massive impression volumes — biggest long-term traffic opportunity",
        "items": [
            {
                "priority": "2a",
                "title": "What is an ERP System? A Beginner's Guide",
                "status": "done",
                "why": "Biggest single opportunity. This needs to be a broad educational piece rather than a product page.",
                "keywords": [
                    "what is an ERP system",
                    "ERP systems UK",
                    "understanding ERP systems",
                    "what is an ERP system and how does it work",
                    "ERP system",
                ],
                "sourcePaths": ["/post/what-is-an-erp-system-a-beginner-s-guide-for-uk-businesses"],
            },
            {
                "priority": "2b",
                "title": "Best ERP for Manufacturers",
                "status": "done",
                "why": "Manufacturing is E3's core vertical. This should stay definitive and comparison-led.",
                "keywords": [
                    "best ERP for manufacturers",
                    "best ERP for manufacturing",
                    "ERP for manufacturers",
                    "manufacturing ERP",
                    "NetSuite for manufacturing",
                ],
                "sourcePaths": [
                    "/post/which-is-the-best-erp-for-manufacturers",
                    "/post/netsuite-for-manufacturing-companies",
                    "/manufacturing",
                ],
            },
            {
                "priority": "2c",
                "title": "Top Benefits of ERP Systems",
                "status": "done",
                "why": "Classic top-of-funnel SEO content with room for click growth once rankings improve.",
                "keywords": [
                    "benefits of ERP systems",
                    "top benefits of ERP",
                    "advantages of ERP",
                    "ERP system benefits",
                ],
                "sourcePaths": ["/post/top-benefits-of-erp-systems-for-your-business"],
            },
        ],
    },
    {
        "id": 3,
        "label": "Phase 3",
        "title": "Strategic Content — Sells E3",
        "subtitle": "Lower volume but captures high-intent visitors actively looking for help",
        "items": [
            {
                "priority": "3a",
                "title": "The Value of Investing in a NetSuite Partner",
                "status": "done",
                "why": "Directly sells E3's service. Explain what a good partner looks like and why E3 fits.",
                "keywords": [
                    "NetSuite partner",
                    "value of NetSuite partner",
                    "why work with a NetSuite partner",
                ],
                "sourcePaths": ["/post/the-value-of-investing-in-a-netsuite-partner"],
            },
            {
                "priority": "3b",
                "title": "10 Signs of a Poor NetSuite Implementation",
                "status": "in_progress",
                "why": "Problem-aware content. People searching this are frustrated, which makes it a strong lead-in to support services.",
                "keywords": [
                    "poor NetSuite implementation",
                    "failed NetSuite implementation",
                    "NetSuite implementation problems",
                ],
                "sourcePaths": ["/post/10-signs-of-a-poor-netsuite-implementation"],
            },
            {
                "priority": "3c",
                "title": "How to Choose the Right ERP Consultant",
                "status": "done",
                "why": "Positions E3 as the benchmark for what a good consultant looks like.",
                "keywords": [
                    "how to choose an ERP consultant",
                    "ERP consultant",
                    "ERP consulting",
                ],
                "sourcePaths": ["/post/how-to-choose-the-right-erp-consultant-for-your-business"],
            },
            {
                "priority": "3d",
                "title": "Why NetSuite Is the Best Choice of Accounting Software",
                "status": "done",
                "why": "Targets finance decision-makers specifically.",
                "keywords": [
                    "NetSuite accounting software",
                    "best accounting software",
                    "NetSuite vs accounting software",
                ],
                "sourcePaths": ["/post/why-netsuite-is-the-best-choice-of-accounting-software"],
            },
        ],
    },
    {
        "id": 4,
        "label": "Phase 4",
        "title": "Fill the Gaps",
        "subtitle": "Write over time as earlier content starts ranking",
        "items": [
            {
                "priority": "4a",
                "title": "Understanding the Role of ERP Systems",
                "status": "done",
                "why": "Could merge into the broader ERP explainer or stand alone.",
                "keywords": [],
                "sourcePaths": ["/post/understanding-the-role-of-erp-systems"],
            },
            {
                "priority": "4b",
                "title": "NetSuite Project Management & Financials",
                "status": "done",
                "why": "Niche, but relevant for service-led buyers.",
                "keywords": [],
                "sourcePaths": ["/post/netsuite-s-unified-project-management-and-financials-improves-project-reporting"],
            },
            {
                "priority": "4c",
                "title": "NetSuite Apps & Extensions",
                "status": "done",
                "why": "Useful for showing the wider NetSuite ecosystem.",
                "keywords": [],
                "sourcePaths": ["/post/a-range-of-applications-that-extend-and-enhance-netsuite"],
            },
            {
                "priority": "4d",
                "title": "7 Warning Signs Your ERP is Holding You Back",
                "status": "todo",
                "why": "Overlaps with the existing article, so this is more about strengthening and redirecting well.",
                "keywords": [],
                "sourcePaths": ["/post/7-warning-signs-your-erp-system-is-holding-you-back"],
            },
            {
                "priority": "4e",
                "title": "Why NetSuite Aftercare is Essential",
                "status": "todo",
                "why": "Feeds directly into the support proposition.",
                "keywords": [],
                "sourcePaths": ["/post/why-netsuite-aftercare-is-essential-for-growing-businesses"],
            },
            {
                "priority": "4f",
                "title": "Optimise End-of-Year Accounting with NetSuite",
                "status": "todo",
                "why": "Seasonal topic worth publishing before year-end demand spikes.",
                "keywords": [],
                "sourcePaths": ["/post/optimise-end-of-year-accounting-with-netsuite"],
            },
        ],
    },
]


def parse_date(value: str, fmt: str) -> date:
    return datetime.strptime(value, fmt).date()


def retry_call(label: str, fn, attempts: int = 3, delay_seconds: int = 3):
    last_exc = None
    for attempt in range(1, attempts + 1):
        try:
            return fn()
        except Exception as exc:
            last_exc = exc
            if attempt < attempts:
                print(
                    f"Warning: {label} failed attempt {attempt}/{attempts}. "
                    f"Retrying in {delay_seconds}s. ({exc})"
                )
                time.sleep(delay_seconds)
    raise last_exc


def latest_reporting_friday(today: date | None = None) -> date:
    current = today or date.today()
    days_since_friday = (current.weekday() - 4) % 7
    return current - timedelta(days=days_since_friday)


def normalise_key(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def to_number(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).strip()
    if not cleaned:
        return None
    cleaned = cleaned.replace(",", "").replace("£", "").replace("$", "")
    cleaned = cleaned.replace("(", "-").replace(")", "")
    cleaned = cleaned.replace(" ", "")
    if cleaned.endswith("%"):
        cleaned = cleaned[:-1]
    try:
        return float(cleaned)
    except ValueError:
        return None


def to_percent(value: object) -> float | None:
    number = to_number(value)
    if number is None:
        return None
    # Handle decimal fractions from exports (e.g. 0.43 -> 43%)
    if 0 <= number <= 1 and "%" not in str(value):
        return round(number * 100, 2)
    return round(number, 2)


def merge_email_metrics(defaults: dict, partial: dict | None) -> dict:
    merged = dict(defaults)
    if not partial:
        return merged
    for key in ("campaignsSent", "totalRecipients", "openRate", "clickRate", "unsubscribes"):
        if key in partial and partial[key] is not None:
            merged[key] = partial[key]
    return merged


def extract_email_metrics_from_mapping(mapping: dict[object, object]) -> tuple[dict | None, int]:
    normalised = {normalise_key(key): value for key, value in mapping.items() if str(key).strip()}
    parsed: dict[str, object] = {}
    score = 0

    for target, aliases in EMAIL_FIELD_ALIASES.items():
        value = None
        for alias in aliases:
            if alias in normalised and str(normalised[alias]).strip():
                value = normalised[alias]
                break
        if value is None:
            continue
        if target in ("campaignsSent", "totalRecipients", "unsubscribes"):
            number = to_number(value)
            if number is None:
                continue
            parsed[target] = int(round(number))
            score += 1
        else:
            number = to_percent(value)
            if number is None:
                continue
            parsed[target] = number
            score += 1

    if score < 3:
        return None, score
    return parsed, score


def parse_email_csv(path: Path) -> dict | None:
    best: dict | None = None
    best_score = 0

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    if not rows:
        return None

    # Tabular (header row + data rows)
    header = rows[0]
    if len(header) >= 2:
        for row in rows[1:]:
            if not any(str(cell).strip() for cell in row):
                continue
            mapping = {header[i]: row[i] if i < len(row) else "" for i in range(len(header))}
            parsed, score = extract_email_metrics_from_mapping(mapping)
            if parsed and score > best_score:
                best = parsed
                best_score = score

    # Key/value layout (label,value)
    kv_mapping = {}
    for row in rows:
        if len(row) < 2:
            continue
        left = str(row[0]).strip()
        right = str(row[1]).strip()
        if left and right:
            kv_mapping[left] = right
    if kv_mapping:
        parsed, score = extract_email_metrics_from_mapping(kv_mapping)
        if parsed and score > best_score:
            best = parsed
            best_score = score

    return best


def parse_email_json(path: Path) -> dict | None:
    with path.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)
    if isinstance(payload, list):
        best = None
        best_score = 0
        for item in payload:
            if not isinstance(item, dict):
                continue
            parsed, score = extract_email_metrics_from_mapping(item)
            if parsed and score > best_score:
                best = parsed
                best_score = score
        return best
    if isinstance(payload, dict):
        parsed, _ = extract_email_metrics_from_mapping(payload)
        return parsed
    return None


def parse_email_xlsx(path: Path) -> dict | None:
    workbook = openpyxl.load_workbook(path, data_only=True)
    best = None
    best_score = 0

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        # Table-style parse using first non-empty row as header
        header_index = None
        for idx, row in enumerate(rows):
            cells = [cell for cell in row if str(cell or "").strip()]
            if len(cells) >= 2:
                header_index = idx
                break
        if header_index is not None:
            header = rows[header_index]
            for row in rows[header_index + 1 :]:
                if not any(str(cell or "").strip() for cell in row):
                    continue
                mapping = {
                    header[i]: row[i] if i < len(row) else None for i in range(len(header))
                }
                parsed, score = extract_email_metrics_from_mapping(mapping)
                if parsed and score > best_score:
                    best = parsed
                    best_score = score

        # Key/value parse from first two columns
        kv_mapping = {}
        for row in rows:
            if len(row) < 2:
                continue
            left = str(row[0] or "").strip()
            right = row[1]
            if left and str(right or "").strip():
                kv_mapping[left] = right
        if kv_mapping:
            parsed, score = extract_email_metrics_from_mapping(kv_mapping)
            if parsed and score > best_score:
                best = parsed
                best_score = score

    return best


def parse_email_file(path: Path) -> dict | None:
    try:
        suffix = path.suffix.lower()
        if suffix == ".csv":
            return parse_email_csv(path)
        if suffix == ".json":
            return parse_email_json(path)
        if suffix == ".xlsx":
            return parse_email_xlsx(path)
    except Exception as exc:
        print(f"Warning: Could not parse email stats from {path.name}: {exc}")
    return None


def score_email_filename(path: Path) -> int:
    name = path.name.lower()
    score = 0
    if any(hint in name for hint in EMAIL_FILENAME_HINTS):
        score += 3
    if "report" in name or "stats" in name or "weekly" in name:
        score += 1
    return score


def iter_candidate_email_files() -> list[Path]:
    if not DOWNLOADS_DIR.exists():
        return []
    candidates = []
    for path in DOWNLOADS_DIR.iterdir():
        if not path.is_file():
            continue
        if path.suffix.lower() not in EMAIL_FILE_EXTENSIONS:
            continue
        if LINKEDIN_EXPORT_RE.match(path.name):
            continue
        candidates.append(path)
    candidates.sort(
        key=lambda p: (score_email_filename(p), p.stat().st_mtime),
        reverse=True,
    )
    return candidates


def latest_email_stats_file(min_mtime: float | None = None) -> tuple[Path, dict] | None:
    for path in iter_candidate_email_files():
        if min_mtime and path.stat().st_mtime < min_mtime:
            continue
        parsed = parse_email_file(path)
        if parsed:
            return path, parsed
    return None


def poll_for_email_stats(start_time: float, wait_minutes: int = EMAIL_WAIT_MINUTES) -> tuple[Path, dict] | None:
    # Use any already-available parseable file first.
    existing = latest_email_stats_file()
    if existing:
        return existing

    deadline = start_time + (max(0, wait_minutes) * 60)
    while time.time() < deadline:
        match = latest_email_stats_file(min_mtime=start_time)
        if match:
            return match
        remaining = int(max(0, deadline - time.time()))
        print(
            f"Waiting for email stats file in {DOWNLOADS_DIR} "
            f"(~{remaining}s remaining)..."
        )
        time.sleep(EMAIL_POLL_SECONDS)

    return None


class GA4Client:
    def __init__(self, service_account_path: Path, property_id: str) -> None:
        scopes = ["https://www.googleapis.com/auth/analytics.readonly"]
        self.credentials = service_account.Credentials.from_service_account_file(
            service_account_path, scopes=scopes
        )
        self.credentials.refresh(Request())
        self.base_url = (
            f"https://analyticsdata.googleapis.com/v1beta/properties/{property_id}:runReport"
        )

    def run_report(
        self,
        *,
        start_date: date,
        end_date: date,
        metrics: list[str],
        dimensions: list[str] | None = None,
        dimension_filter: dict | None = None,
        order_bys: list[dict] | None = None,
        limit: int | None = None,
    ) -> list[dict]:
        payload: dict = {
            "dateRanges": [
                {"startDate": start_date.isoformat(), "endDate": end_date.isoformat()}
            ],
            "metrics": [{"name": name} for name in metrics],
        }
        if dimensions:
            payload["dimensions"] = [{"name": name} for name in dimensions]
        if dimension_filter:
            payload["dimensionFilter"] = dimension_filter
        if order_bys:
            payload["orderBys"] = order_bys
        if limit:
            payload["limit"] = limit

        response = requests.post(
            self.base_url,
            headers={"Authorization": f"Bearer {self.credentials.token}"},
            json=payload,
            timeout=30,
        )
        response.raise_for_status()
        body = response.json()
        rows = []
        for row in body.get("rows", []):
            rows.append(
                {
                    "dimensions": [value["value"] for value in row.get("dimensionValues", [])],
                    "metrics": [value["value"] for value in row.get("metricValues", [])],
                }
            )
        return rows


class SearchConsoleClient:
    def __init__(self, service_account_path: Path) -> None:
        scopes = ["https://www.googleapis.com/auth/webmasters.readonly"]
        self.credentials = service_account.Credentials.from_service_account_file(
            service_account_path, scopes=scopes
        )
        self.credentials.refresh(Request())

    def query(
        self,
        *,
        site_url: str,
        start_date: date,
        end_date: date,
        dimensions: list[str] | None = None,
        row_limit: int = 10,
    ) -> dict:
        encoded_site = urllib.parse.quote(site_url, safe="")
        url = (
            f"https://www.googleapis.com/webmasters/v3/sites/{encoded_site}/searchAnalytics/query"
        )
        payload: dict = {
            "startDate": start_date.isoformat(),
            "endDate": end_date.isoformat(),
            "rowLimit": row_limit,
        }
        if dimensions:
            payload["dimensions"] = dimensions
        response = requests.post(
            url,
            headers={"Authorization": f"Bearer {self.credentials.token}"},
            json=payload,
            timeout=30,
        )
        response.raise_for_status()
        return response.json()


def latest_linkedin_weekly_export() -> Path:
    candidates = []
    for directory in (DOWNLOADS_DIR, PLAYWRIGHT_DOWNLOADS_DIR, EXPORTS_DIR):
        if not directory.exists():
            continue
        for path in directory.iterdir():
            match = LINKEDIN_EXPORT_RE.match(path.name)
            if match:
                end_date = parse_date(match.group(2), "%Y-%m-%d")
                candidates.append((end_date, path.stat().st_mtime, path))
    if not candidates:
        raise FileNotFoundError(
            "No weekly LinkedIn workbook found in Downloads, .playwright-cli, or src/exports."
        )
    candidates.sort(reverse=True)
    return candidates[0][2]


def latest_linkedin_audience_export() -> Path | None:
    candidates = []
    for directory in (DOWNLOADS_DIR, PLAYWRIGHT_DOWNLOADS_DIR, EXPORTS_DIR):
        if not directory.exists():
            continue
        for path in directory.iterdir():
            match = LINKEDIN_AUDIENCE_EXPORT_RE.match(path.name)
            if not match:
                continue
            end_date = parse_date(match.group(2), "%Y-%m-%d")
            candidates.append((end_date, path.stat().st_mtime, path))
    if not candidates:
        return None
    candidates.sort(reverse=True)
    return candidates[0][2]


def latest_linkedin_annual_export() -> Path | None:
    candidates = []
    for directory in (DOWNLOADS_DIR, PLAYWRIGHT_DOWNLOADS_DIR, EXPORTS_DIR):
        if not directory.exists():
            continue
        for path in directory.iterdir():
            match = LINKEDIN_EXPORT_RE.match(path.name)
            if not match:
                continue
            start_date = parse_date(match.group(1), "%Y-%m-%d")
            end_date = parse_date(match.group(2), "%Y-%m-%d")
            span_days = (end_date - start_date).days
            if span_days < 300:
                continue
            candidates.append((end_date, path.stat().st_mtime, path))
    if not candidates:
        return None
    candidates.sort(reverse=True)
    return candidates[0][2]


def copy_export(source: Path) -> None:
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    target = EXPORTS_DIR / source.name
    if source.resolve() != target.resolve():
        shutil.copy2(source, target)


def to_int(value: str | float | int) -> int:
    return int(round(float(value)))


def to_float(value: str | float | int, digits: int = 1) -> float:
    return round(float(value), digits)


def normalise_path(path: str) -> str | None:
    if not path:
        return None
    path = path.strip()
    if any(path == prefix or path.startswith(f"{prefix}/") for prefix in INTERNAL_PATH_PREFIXES):
        return None
    return path


def normalise_page_title(title: str) -> str | None:
    if not title:
        return None
    if title in TITLE_ALIASES:
        return TITLE_ALIASES[title]
    if "Internal" in title or title.startswith("Changelog") or title.startswith("Page Not Found"):
        return None
    trimmed = title.replace(" | ERP Experts", "").strip()
    if trimmed.startswith("SEO Content Roadmap"):
        return None
    return trimmed


def humanise_cta_label(label: str) -> str:
    if label in CTA_LABEL_ALIASES:
        return CTA_LABEL_ALIASES[label]
    text = label.replace("_", " ").strip().title()
    text = re.sub(r"\bCta\b", "CTA", text)
    text = re.sub(r"\bNetscore\b", "NETscore", text)
    text = re.sub(r"\bFaq\b", "FAQ", text)
    text = re.sub(r"\bSeo\b", "SEO", text)
    text = re.sub(r"\bAi\b", "AI", text)
    return text


def aggregate_weekly_linkedin(workbook_path: Path) -> dict:
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    engagement = workbook["ENGAGEMENT"]
    followers = workbook["FOLLOWERS"]

    week_end_match = re.search(r"_(\d{4}-\d{2}-\d{2})_RicWilson", workbook_path.name)
    if not week_end_match:
        raise ValueError(f"Could not parse week ending from {workbook_path.name}")
    week_ending = parse_date(week_end_match.group(1), "%Y-%m-%d")
    week_start = week_ending - timedelta(days=6)

    impressions = 0
    engagements = 0
    for row in engagement.iter_rows(min_row=2, values_only=True):
        date_str, day_impressions, day_engagements = row
        if not date_str:
            continue
        day = parse_date(date_str, "%m/%d/%Y")
        if week_start <= day <= week_ending:
            impressions += int(day_impressions or 0)
            engagements += int(day_engagements or 0)

    total_followers = int(followers["B1"].value or 0)
    new_followers = 0
    for row in followers.iter_rows(min_row=4, values_only=True):
        date_str, daily_new = row
        if not date_str:
            continue
        day = parse_date(date_str, "%m/%d/%Y")
        if week_start <= day <= week_ending:
            new_followers += int(daily_new or 0)

    engagement_rate = round((engagements / impressions) * 100, 1) if impressions else 0.0
    return {
        "weekEnding": week_ending.isoformat(),
        "linkedin": {
            "followers": total_followers,
            "impressions": impressions,
            "engagementRate": engagement_rate,
            "engagements": engagements,
            "newFollowers": new_followers,
        },
    }


def parse_linkedin_year(workbook_path: Path) -> dict:
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    discovery = workbook["DISCOVERY"]
    demographics_sheet = workbook["DEMOGRAPHICS"]

    period_text = str(discovery["B1"].value or "").strip()
    period_match = re.match(r"(\d{1,2}/\d{1,2}/\d{4})\s*-\s*(\d{1,2}/\d{1,2}/\d{4})", period_text)
    if not period_match:
        raise ValueError(f"Could not parse LinkedIn year period from {workbook_path.name}")

    categories = {
        "jobTitles": [],
        "locations": [],
        "industries": [],
        "seniority": [],
        "companySize": [],
    }
    category_map = {
        "Job titles": ("jobTitles", "title"),
        "Locations": ("locations", "location"),
        "Industries": ("industries", "industry"),
        "Seniority": ("seniority", "level"),
        "Company size": ("companySize", "size"),
    }

    for category, value, pct in demographics_sheet.iter_rows(min_row=2, values_only=True):
        if category not in category_map or not value or pct is None:
            continue
        key, field = category_map[category]
        label = str(value).strip()
        if key == "companySize" and "employees" not in label:
            label = f"{label} employees"
        categories[key].append(
            {
                field: label,
                "pct": round(float(pct) * 100, 1),
            }
        )

    return {
        "start": parse_date(period_match.group(1), "%m/%d/%Y").isoformat(),
        "end": parse_date(period_match.group(2), "%m/%d/%Y").isoformat(),
        "totalImpressions": to_int(discovery["B2"].value or 0),
        "membersReached": to_int(discovery["B3"].value or 0),
        "demographics": categories,
    }


def parse_linkedin_top_posts(workbook_path: Path) -> list[dict]:
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    sheet = workbook["TOP POSTS"]

    impressions_by_url: dict[str, int] = {}
    posts: list[dict] = []

    for row in sheet.iter_rows(min_row=4, values_only=True):
        eng_url, eng_date, engagements, _, imp_url, imp_date, impressions, _ = row

        if imp_url and impressions is not None:
            impressions_by_url[str(imp_url).strip()] = to_int(impressions)

        if not eng_url or engagements is None:
            continue

        url = str(eng_url).strip()
        if not url:
            continue

        post_date = parse_date(str(eng_date).strip(), "%m/%d/%Y").isoformat() if eng_date else None
        posts.append(
            {
                "url": url,
                "date": post_date,
                "engagements": to_int(engagements),
            }
        )

    for post in posts:
        impressions = impressions_by_url.get(post["url"], 0)
        engagements = post["engagements"]
        post["impressions"] = impressions
        post["engagementRate"] = round((engagements / impressions) * 100, 2) if impressions else 0.0
        post["excerpt"] = "Open post on LinkedIn"
        post["votes"] = 0
        post["comments"] = 0
        post["reposts"] = 0

    posts.sort(key=lambda item: item["engagementRate"], reverse=True)
    return posts[:10]


def validate_audience_export(workbook_path: Path, min_end_date: date) -> tuple[bool, str]:
    match = LINKEDIN_AUDIENCE_EXPORT_RE.match(workbook_path.name)
    if not match:
        return False, f"Invalid audience export filename: {workbook_path.name}"
    end_date = parse_date(match.group(2), "%Y-%m-%d")
    if end_date < min_end_date:
        return False, (
            f"Audience export is stale ({end_date.isoformat()}); "
            f"needs at least {min_end_date.isoformat()}"
        )
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    sheet_names = set(workbook.sheetnames)
    required = {"FOLLOWERS", "DEMOGRAPHICS"}
    if not required.issubset(sheet_names):
        return False, (
            f"Audience export missing required sheets: {sorted(required - sheet_names)}"
        )
    return True, f"Audience export OK ({workbook_path.name})"


def refresh_linkedin_payload(
    reports: dict, *, strict: bool
) -> tuple[dict, dict, list[dict], list[str]]:
    notes: list[str] = []
    linkedin_export = latest_linkedin_weekly_export()
    copy_export(linkedin_export)
    linkedin_week = aggregate_weekly_linkedin(linkedin_export)
    linkedin_year_export = latest_linkedin_annual_export()
    linkedin_year = (
        parse_linkedin_year(linkedin_year_export)
        if linkedin_year_export
        else reports.get("linkedinYear", {})
    )
    top_linkedin_posts = (
        parse_linkedin_top_posts(linkedin_year_export)
        if linkedin_year_export
        else reports.get("topLinkedInPosts", [])
    )

    linkedin_week_ending = parse_date(linkedin_week["weekEnding"], "%Y-%m-%d")
    latest_report_date = latest_reporting_friday()
    if linkedin_week_ending < latest_report_date:
        msg = (
            "Fresh LinkedIn weekly export not available. "
            f"Using fallback workbook through {linkedin_week_ending.isoformat()} "
            f"for report week ending {latest_report_date.isoformat()}."
        )
        notes.append(msg)
        if strict:
            raise RuntimeError(msg)

    audience_export = latest_linkedin_audience_export()
    if audience_export is None:
        msg = (
            "No LinkedIn Audience export found. Expected an Audience_*_RicWilson workbook "
            "with FOLLOWERS and DEMOGRAPHICS sheets."
        )
        notes.append(msg)
        if strict:
            raise RuntimeError(msg)
    else:
        min_end_date = latest_report_date - timedelta(days=2)
        ok, message = validate_audience_export(audience_export, min_end_date=min_end_date)
        notes.append(message)
        if strict and not ok:
            raise RuntimeError(message)

    return linkedin_week, linkedin_year, top_linkedin_posts, notes


def bucket_sources(rows: list[dict], total_sessions: int, percent: bool) -> dict:
    buckets = {"organic": 0, "direct": 0, "referral": 0, "social": 0, "unassigned": 0}
    for row in rows:
        channel = row["dimensions"][0]
        sessions = to_int(row["metrics"][0])
        if channel == "Direct":
            buckets["direct"] += sessions
        elif channel == "Organic Search":
            buckets["organic"] += sessions
        elif channel == "Referral":
            buckets["referral"] += sessions
        elif "Social" in channel:
            buckets["social"] += sessions
        else:
            buckets["unassigned"] += sessions

    if not percent:
        return {key: value for key, value in buckets.items() if value > 0}

    if total_sessions <= 0:
        return {key: 0 for key in buckets}
    return {
        key: round((value / total_sessions) * 100) if value > 0 else 0
        for key, value in buckets.items()
    }


def build_page_engagement(rows: list[dict]) -> list[dict]:
    pages = []
    seen = set()
    for row in rows:
        title = normalise_page_title(row["dimensions"][0])
        if not title or title in seen:
            continue
        seen.add(title)
        views = to_int(row["metrics"][0])
        sessions = to_int(row["metrics"][1])
        bounce_rate = round(float(row["metrics"][4]) * 100, 1)
        pages.append(
            {
                "page": title,
                "avgTimeOnPage": to_int(row["metrics"][3]),
                "views": views,
                "sessions": sessions,
                "activeUsers": to_int(row["metrics"][2]),
                "bounceRate": bounce_rate,
            }
        )
    return pages[:9]


def build_exit_pages(page_engagement: list[dict]) -> list[dict]:
    exit_pages = []
    for page in page_engagement:
        exits = round(page["sessions"] * (page["bounceRate"] / 100))
        exit_pages.append(
            {
                "page": page["page"],
                "exitRate": page["bounceRate"],
                "exits": exits,
            }
        )
    exit_pages.sort(key=lambda item: item["exitRate"], reverse=True)
    return exit_pages


def normalise_search_console_path(url: str) -> str | None:
    if not url:
        return None
    parsed = urlparse(url)
    path = parsed.path or "/"
    if parsed.query:
        path = f"{path}?{parsed.query}"
    return path


def to_percent_string_number(value: str) -> float:
    cleaned = value.strip().replace("%", "")
    return round(float(cleaned), 2)


def weighted_avg_position(rows: list[dict]) -> float:
    if not rows:
        return 0.0
    weighted_impressions = sum(row["impressions"] for row in rows)
    if weighted_impressions <= 0:
        return round(sum(row["avgPosition"] for row in rows) / len(rows), 2)
    return round(
        sum(row["avgPosition"] * row["impressions"] for row in rows) / weighted_impressions,
        2,
    )


def build_roadmap_phases(all_pages: list[dict]) -> list[dict]:
    page_lookup = {page["path"]: page for page in all_pages}
    phases = []

    for phase in ROADMAP_PHASES_TEMPLATE:
        phase_items = []
        for item in phase["items"]:
            matched_rows = [
                page_lookup[path]
                for path in item.get("sourcePaths", [])
                if path in page_lookup
            ]
            phase_items.append(
                {
                    "priority": item["priority"],
                    "title": item["title"],
                    "impressions": sum(row["impressions"] for row in matched_rows),
                    "position": weighted_avg_position(matched_rows),
                    "clicks": sum(row["clicks"] for row in matched_rows),
                    "status": item["status"],
                    "why": item["why"],
                    "keywords": item["keywords"],
                }
            )
        phases.append(
            {
                "id": phase["id"],
                "label": phase["label"],
                "title": phase["title"],
                "subtitle": phase["subtitle"],
                "items": phase_items,
            }
        )

    return phases


def build_search_demand_signals(query_rows: list[dict], *, max_items: int = 12) -> dict:
    """Extract ERP/NetSuite demand and likely content gaps from Search Console query rows."""
    watched_terms = ("erp", "netsuite")
    brand_terms = (
        "erp experts",
        "erpexperts",
        "erp experts uk",
        "erp experts ltd",
    )
    filtered: list[dict] = []

    for row in query_rows:
        query = (row.get("query") or "").strip().lower()
        if not query:
            continue
        if not any(term in query for term in watched_terms):
            continue
        if any(term in query for term in brand_terms):
            continue
        filtered.append(row)

    # Prioritise demand where impressions exist but clicks lag and ranking is not yet strong.
    filtered.sort(
        key=lambda item: (
            item.get("impressions", 0),
            -item.get("clicks", 0),
            -item.get("position", 0),
        ),
        reverse=True,
    )

    top_queries = [
        {
            "query": item.get("query", ""),
            "impressions": int(round(item.get("impressions", 0) or 0)),
            "clicks": int(round(item.get("clicks", 0) or 0)),
            "ctr": round(float(item.get("ctr", 0) or 0) * 100, 2),
            "position": round(float(item.get("position", 0) or 0), 2),
            "landingPage": item.get("page", ""),
        }
        for item in filtered[:max_items]
    ]

    erp_pattern = re.compile(r"\berp\s+(?:in|for)\s+([a-z0-9&\-\s]{2,60})\b", re.IGNORECASE)
    sector_counts: dict[str, dict[str, float]] = {}
    for item in filtered:
        query = item.get("query", "")
        match = erp_pattern.search(query)
        if not match:
            continue
        sector = re.sub(r"\s+", " ", match.group(1)).strip(" -")
        if not sector:
            continue
        key = sector.lower()
        entry = sector_counts.setdefault(
            key,
            {"term": sector, "impressions": 0.0, "clicks": 0.0, "queries": 0},
        )
        entry["impressions"] += float(item.get("impressions", 0) or 0)
        entry["clicks"] += float(item.get("clicks", 0) or 0)
        entry["queries"] += 1

    erp_in_candidates = sorted(
        [
            {
                "term": value["term"],
                "impressions": int(round(value["impressions"])),
                "clicks": int(round(value["clicks"])),
                "queryCount": int(value["queries"]),
            }
            for value in sector_counts.values()
        ],
        key=lambda x: (x["impressions"], -x["clicks"], x["queryCount"]),
        reverse=True,
    )[:8]

    # Gap signal: meaningful demand but weak traction.
    content_gaps = [
        q
        for q in top_queries
        if q["impressions"] >= 20 and (q["clicks"] <= 1 or q["ctr"] < 1.0) and q["position"] >= 8
    ][:10]

    def classify_intent(query: str) -> str:
        q = query.lower()
        if " vs " in q or "versus" in q or "compare" in q:
            return "comparison"
        if any(token in q for token in ("migrate", "migration", "switch", "replace")):
            return "migration"
        if any(token in q for token in ("consultant", "partner", "implementation", "agency")):
            return "implementation"
        if any(token in q for token in ("cost", "price", "pricing", "tco", "roi")):
            return "commercial"
        if re.search(r"\berp\s+(?:in|for)\b", q):
            return "vertical"
        return "discovery"

    def topic_title(query: str, intent: str) -> str:
        clean = query.strip().rstrip("?")
        if intent == "comparison":
            return f"NetSuite {clean.title()} Guide"
        if intent == "migration":
            return f"{clean.title()}: Migration Playbook"
        if intent == "implementation":
            return f"{clean.title()}: How to Choose and Deliver"
        if intent == "commercial":
            return f"{clean.title()}: Cost and Value Breakdown"
        if intent == "vertical":
            return f"{clean.title()}: Industry Guide"
        return f"{clean.title()}: Practical Guide"

    def score_topic(item: dict) -> int:
        impressions = int(item.get("impressions", 0) or 0)
        clicks = int(item.get("clicks", 0) or 0)
        ctr = float(item.get("ctr", 0) or 0)
        position = float(item.get("position", 0) or 0)
        intent = classify_intent(item.get("query", ""))
        score = 0
        score += min(45, impressions // 5)
        if clicks <= 1:
            score += 18
        if ctr < 1.0:
            score += 12
        elif ctr < 2.0:
            score += 6
        if position >= 8:
            score += 12
        elif position >= 5:
            score += 6
        if intent in {"comparison", "migration", "implementation"}:
            score += 10
        elif intent == "commercial":
            score += 8
        return max(1, min(100, score))

    topic_backlog = []
    seen_queries = set()
    for item in top_queries:
        query = (item.get("query") or "").strip()
        if not query:
            continue
        query_key = query.lower()
        if query_key in seen_queries:
            continue
        seen_queries.add(query_key)
        intent = classify_intent(query)
        score = score_topic(item)
        topic_backlog.append(
            {
                "query": query,
                "intent": intent,
                "score": score,
                "suggestedTitle": topic_title(query, intent),
                "impressions": item.get("impressions", 0),
                "clicks": item.get("clicks", 0),
                "ctr": item.get("ctr", 0),
                "position": item.get("position", 0),
                "whyNow": "High impressions with weak traction suggests reachable growth from a better-matched page.",
            }
        )

    topic_backlog.sort(key=lambda x: (x["score"], x["impressions"]), reverse=True)
    topic_backlog = topic_backlog[:10]

    return {
        "erpNetsuiteQueries": top_queries,
        "erpInPatterns": erp_in_candidates,
        "contentGaps": content_gaps,
        "topicBacklogV2": topic_backlog,
    }


def load_search_console_fallback(*, start_date: date, end_date: date) -> dict:
    if not GSC_FALLBACK_PAGES_CSV.exists():
        raise FileNotFoundError(f"Missing Search Console fallback CSV: {GSC_FALLBACK_PAGES_CSV}")

    pages = []
    with GSC_FALLBACK_PAGES_CSV.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            raw_url = (row.get("Top pages") or "").strip()
            path = normalise_search_console_path(raw_url)
            if not path:
                continue
            pages.append(
                {
                    "path": path,
                    "impressions": int(round(float((row.get("Impressions") or "0").replace(",", "")))),
                    "clicks": int(round(float((row.get("Clicks") or "0").replace(",", "")))),
                    "ctr": to_percent_string_number(row.get("CTR") or "0"),
                    "avgPosition": round(float((row.get("Position") or "0").replace(",", "")), 2),
                }
            )

    total_impressions = sum(page["impressions"] for page in pages)
    total_clicks = sum(page["clicks"] for page in pages)
    avg_ctr = round((total_clicks / total_impressions) * 100, 2) if total_impressions else 0.0

    return {
        "period": f"{start_date.isoformat()} to {end_date.isoformat()}",
        "totalImpressions": total_impressions,
        "totalClicks": total_clicks,
        "avgCTR": avg_ctr,
        "topLandingPages": pages[:5],
        "allLandingPages": pages,
        "roadmapPhases": build_roadmap_phases(pages),
        "demandSignals": {
            "erpNetsuiteQueries": [],
            "erpInPatterns": [],
            "contentGaps": [],
            "topicBacklogV2": [],
            "note": "Query-level demand signals unavailable in CSV fallback mode.",
        },
        "source": "fallback_csv",
    }


def read_export_with_metadata(path: Path) -> tuple[dict[str, str], list[dict[str, str]]]:
    metadata: dict[str, str] = {}
    header: list[str] | None = None
    rows: list[dict[str, str]] = []

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        for raw_row in reader:
            if not raw_row:
                continue
            first = (raw_row[0] or "").strip()
            if first.startswith("#"):
                line = first.lstrip("#").strip()
                if ":" in line:
                    key, value = line.split(":", 1)
                    metadata[key.strip()] = value.strip()
                elif re.fullmatch(r"\d{8}-\d{8}", line):
                    metadata["Date range"] = line
                continue
            if header is None:
                header = [cell.strip() for cell in raw_row]
                continue
            row = {}
            for index, column in enumerate(header):
                row[column] = raw_row[index].strip() if index < len(raw_row) else ""
            rows.append(row)

    if header is None:
        raise ValueError(f"Could not find CSV header in {path}")

    return metadata, rows


def parse_export_date_range(metadata: dict[str, str]) -> tuple[date, date]:
    if metadata.get("Start date") and metadata.get("End date"):
        return (
            parse_date(metadata["Start date"], "%Y%m%d"),
            parse_date(metadata["End date"], "%Y%m%d"),
        )
    if metadata.get("Date range"):
        start_raw, end_raw = metadata["Date range"].split("-", 1)
        return parse_date(start_raw, "%Y%m%d"), parse_date(end_raw, "%Y%m%d")
    raise ValueError("CSV metadata is missing a parseable date range")


def load_ga4_fallback(reports: dict) -> tuple[dict, dict]:
    required_paths = [
        GA_EVENTS_FALLBACK_CSV,
        GA_CHANNELS_FALLBACK_CSV,
        GA_PAGES_FALLBACK_CSV,
        GA_COUNTRIES_FALLBACK_CSV,
    ]
    for path in required_paths:
        if not path.exists():
            raise FileNotFoundError(f"Missing GA fallback CSV: {path}")

    events_meta, event_rows = read_export_with_metadata(GA_EVENTS_FALLBACK_CSV)
    period_start, period_end = parse_export_date_range(events_meta)
    _, channel_rows = read_export_with_metadata(GA_CHANNELS_FALLBACK_CSV)
    _, page_rows = read_export_with_metadata(GA_PAGES_FALLBACK_CSV)
    _, country_rows = read_export_with_metadata(GA_COUNTRIES_FALLBACK_CSV)

    event_counts = {
        row["Event name"]: to_int(row.get("Event count") or 0)
        for row in event_rows
        if row.get("Event name")
    }
    event_users = {
        row["Event name"]: to_int(row.get("Total users") or 0)
        for row in event_rows
        if row.get("Event name")
    }
    total_sessions = sum(to_int(row.get("Sessions") or 0) for row in channel_rows)
    total_page_views = event_counts.get("page_view", 0)
    total_users = max(event_users.values()) if event_users else 0
    total_new_users = event_counts.get("first_visit", 0)

    period_sources = []
    for row in channel_rows:
        channel = row.get("Session primary channel group (Default channel group)", "")
        sessions = to_int(row.get("Sessions") or 0)
        period_sources.append({"dimensions": [channel], "metrics": [str(sessions)]})

    page_rows_for_build = []
    for row in page_rows:
        title = row.get("Page title and screen class", "")
        if not title:
            continue
        page_rows_for_build.append(
            {
                "dimensions": [title],
                "metrics": [
                    row.get("Views", "0"),
                    row.get("Sessions", "0"),
                    row.get("Active users", "0"),
                    row.get("Average session duration", "0"),
                    row.get("Bounce rate", "0"),
                ],
            }
        )

    page_engagement = build_page_engagement(page_rows_for_build)
    page_lookup = {page["page"]: page for page in page_engagement}

    period_payload = {
        "start": period_start.isoformat(),
        "end": period_end.isoformat(),
        "totalSessions": total_sessions,
        "totalPageViews": total_page_views,
        "totalUsers": total_users,
        "totalNewUsers": total_new_users,
        "totalCTAClicks": event_counts.get("cta_click", 0),
        "totalLeads": event_counts.get("generate_lead", 0),
        "countries": [
            {"country": row["Country"], "users": to_int(row.get("Active users") or 0)}
            for row in country_rows
            if row.get("Country")
        ][:8],
        "trafficSources": bucket_sources(period_sources, total_sessions, percent=False),
        "funnel": [
            {
                "step": "Landed on site",
                "users": total_users,
                "description": "Active users from the latest local GA export",
            },
            {
                "step": "Viewed homepage",
                "users": page_lookup.get("Homepage", {}).get("activeUsers", 0),
                "description": "Viewed the homepage",
            },
            {
                "step": "Visited contact page",
                "users": page_lookup.get("Contact", {}).get("activeUsers", 0),
                "description": "Viewed the contact page",
            },
            {
                "step": "Clicked a CTA",
                "users": event_users.get("cta_click", 0),
                "description": "Triggered a CTA click event",
            },
            {
                "step": "Became a lead",
                "users": event_users.get("generate_lead", 0),
                "description": "Triggered a generate_lead event",
            },
        ],
        "pageBounceRates": [
            {"page": page["page"], "bounceRate": page["bounceRate"]}
            for page in page_engagement
        ],
        "ctaBreakdown": reports.get("ga4Period", {}).get("ctaBreakdown", []),
        "navigationFlow": reports.get("ga4Period", {}).get("navigationFlow", {}),
        "pageEngagement": page_engagement,
        "exitPages": build_exit_pages(page_engagement),
        "seoInsights": reports.get("ga4Period", {}).get("seoInsights", {}),
        "source": "fallback_csv",
    }

    fallback_week = None
    previous_week = reports.get("weeks", [{}])[0]
    if previous_week:
        fallback_week = dict(previous_week.get("ga", {}))
    if fallback_week is None:
        fallback_week = {}

    return period_payload, fallback_week


def refresh_search_console(
    client: SearchConsoleClient, *, start_date: date, end_date: date
) -> dict:
    totals = client.query(
        site_url=SEARCH_CONSOLE_SITE,
        start_date=start_date,
        end_date=end_date,
        row_limit=1,
    )
    top_pages = client.query(
        site_url=SEARCH_CONSOLE_SITE,
        start_date=start_date,
        end_date=end_date,
        dimensions=["page"],
        row_limit=250,
    )
    top_queries = client.query(
        site_url=SEARCH_CONSOLE_SITE,
        start_date=start_date,
        end_date=end_date,
        dimensions=["query", "page"],
        row_limit=1000,
    )

    rows = top_pages.get("rows", [])
    pages = []
    for row in rows:
        page_url = row.get("keys", [""])[0]
        path = normalise_search_console_path(page_url)
        if not path:
            continue
        pages.append(
            {
                "path": path,
                "impressions": round(row.get("impressions", 0)),
                "clicks": round(row.get("clicks", 0)),
                "ctr": round(row.get("ctr", 0) * 100, 2),
                "avgPosition": round(row.get("position", 0), 2),
            }
        )

    query_rows = []
    for row in top_queries.get("rows", []):
        keys = row.get("keys", [])
        query = keys[0] if len(keys) > 0 else ""
        page_url = keys[1] if len(keys) > 1 else ""
        query_rows.append(
            {
                "query": query,
                "page": normalise_search_console_path(page_url) or page_url,
                "impressions": row.get("impressions", 0),
                "clicks": row.get("clicks", 0),
                "ctr": row.get("ctr", 0),
                "position": row.get("position", 0),
            }
        )

    return {
        "period": f"{start_date.isoformat()} to {end_date.isoformat()}",
        "totalImpressions": round(totals.get("rows", [{}])[0].get("impressions", 0)),
        "totalClicks": round(totals.get("rows", [{}])[0].get("clicks", 0)),
        "avgCTR": round(totals.get("rows", [{}])[0].get("ctr", 0) * 100, 2),
        "topLandingPages": pages[:5],
        "allLandingPages": pages,
        "roadmapPhases": build_roadmap_phases(pages),
        "demandSignals": build_search_demand_signals(query_rows),
        "source": "search_console_api",
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Refresh ERP Experts reporting data")
    parser.add_argument(
        "--linkedin-only",
        action="store_true",
        help="Only refresh LinkedIn-derived sections and validate LinkedIn exports",
    )
    parser.add_argument(
        "--strict-linkedin",
        action="store_true",
        help="Fail if weekly Content or Audience exports are missing/stale",
    )
    args = parser.parse_args()

    if not REPORTS_PATH.exists():
        raise FileNotFoundError(f"Missing reports file: {REPORTS_PATH}")
    if not SERVICE_ACCOUNT_PATH.exists():
        raise FileNotFoundError(f"Missing service account JSON: {SERVICE_ACCOUNT_PATH}")

    with REPORTS_PATH.open("r", encoding="utf-8") as handle:
        reports = json.load(handle)

    base_email_defaults = {
        "campaignsSent": 0,
        "totalRecipients": 0,
        "openRate": 0,
        "clickRate": 0,
        "unsubscribes": 0,
    }
    previous_email = (
        reports.get("weeks", [{}])[0].get("email", {})
        if reports.get("weeks")
        else {}
    )
    email_defaults = merge_email_metrics(base_email_defaults, previous_email)

    run_started = time.time()
    email_match = None
    latest_report_date = latest_reporting_friday()
    week_label = latest_report_date.isoformat()

    linkedin_week, linkedin_year, top_linkedin_posts, linkedin_notes = refresh_linkedin_payload(
        reports, strict=args.strict_linkedin
    )
    for note in linkedin_notes:
        print(f"LinkedIn check: {note}")

    if args.linkedin_only:
        reports["lastUpdated"] = date.today().isoformat()
        reports["dataThrough"] = week_label
        reports["linkedinYear"] = linkedin_year
        reports["topLinkedInPosts"] = top_linkedin_posts
        existing_weeks = reports.get("weeks", [])
        current = existing_weeks[0] if existing_weeks else {}
        reports["weeks"] = [
            {
                "weekEnding": week_label,
                "ga": current.get("ga", {}),
                "linkedin": linkedin_week["linkedin"],
                "email": current.get("email", email_defaults),
            }
        ] + [week for week in existing_weeks if week.get("weekEnding") != week_label]

        with REPORTS_PATH.open("w", encoding="utf-8") as handle:
            json.dump(reports, handle, indent=2, ensure_ascii=False)
            handle.write("\n")
        print("Updated reports.json (LinkedIn only)")
        print(
            "  Latest LinkedIn week:",
            linkedin_week["linkedin"]["followers"],
            "followers,",
            linkedin_week["linkedin"]["impressions"],
            "impressions,",
            f"{linkedin_week['linkedin']['engagementRate']}% engagement",
        )
        return

    with ThreadPoolExecutor(max_workers=1) as pool:
        email_future = pool.submit(poll_for_email_stats, run_started, EMAIL_WAIT_MINUTES)

        ga_week_end = latest_report_date - timedelta(days=1)
        ga_week_start = ga_week_end - timedelta(days=6)
        period_end = ga_week_end
        period_start = period_end - timedelta(days=27)

        try:
            search_console_client = retry_call(
                "Search Console client initialisation",
                lambda: SearchConsoleClient(SERVICE_ACCOUNT_PATH),
            )
        except Exception as exc:
            print(
                "Warning: Search Console client initialisation failed. "
                f"Keeping previous SEO insights. ({exc})"
            )
            search_console_client = None

        try:
            ga_client = retry_call(
                "GA4 client initialisation",
                lambda: GA4Client(SERVICE_ACCOUNT_PATH, GA4_PROPERTY_ID),
            )

            weekly_overall = retry_call(
                "GA4 weekly overall report",
                lambda: ga_client.run_report(
                    start_date=ga_week_start,
                    end_date=ga_week_end,
                    metrics=[
                        "sessions",
                        "screenPageViews",
                        "activeUsers",
                        "newUsers",
                        "averageSessionDuration",
                    ],
                ),
            )[0]
            weekly_events = retry_call(
                "GA4 weekly events report",
                lambda: ga_client.run_report(
                    start_date=ga_week_start,
                    end_date=ga_week_end,
                    dimensions=["eventName"],
                    metrics=["eventCount", "totalUsers"],
                    order_bys=[{"metric": {"metricName": "eventCount"}, "desc": True}],
                    limit=20,
                ),
            )
            weekly_sources = retry_call(
                "GA4 weekly sources report",
                lambda: ga_client.run_report(
                    start_date=ga_week_start,
                    end_date=ga_week_end,
                    dimensions=["sessionDefaultChannelGroup"],
                    metrics=["sessions"],
                    order_bys=[{"metric": {"metricName": "sessions"}, "desc": True}],
                    limit=10,
                ),
            )
            weekly_top_pages_raw = retry_call(
                "GA4 weekly top pages report",
                lambda: ga_client.run_report(
                    start_date=ga_week_start,
                    end_date=ga_week_end,
                    dimensions=["pagePath"],
                    metrics=["screenPageViews"],
                    order_bys=[{"metric": {"metricName": "screenPageViews"}, "desc": True}],
                    limit=20,
                ),
            )

            weekly_top_pages = []
            for row in weekly_top_pages_raw:
                path = normalise_path(row["dimensions"][0])
                if not path:
                    continue
                weekly_top_pages.append({"path": path, "views": to_int(row["metrics"][0])})
                if len(weekly_top_pages) == 8:
                    break

            weekly_event_counts = {
                row["dimensions"][0]: to_int(row["metrics"][0]) for row in weekly_events
            }
            weekly_ga = {
                "sessions": to_int(weekly_overall["metrics"][0]),
                "pageViews": to_int(weekly_overall["metrics"][1]),
                "users": to_int(weekly_overall["metrics"][2]),
                "newUsers": to_int(weekly_overall["metrics"][3]),
                "avgSessionDuration": to_int(weekly_overall["metrics"][4]),
                "topPages": weekly_top_pages,
                "trafficSources": bucket_sources(
                    weekly_sources, to_int(weekly_overall["metrics"][0]), percent=True
                ),
                "ctaClicks": weekly_event_counts.get("cta_click", 0),
                "leads": weekly_event_counts.get("generate_lead", 0),
            }

            period_overall = retry_call(
                "GA4 period overall report",
                lambda: ga_client.run_report(
                    start_date=period_start,
                    end_date=period_end,
                    metrics=["sessions", "screenPageViews", "activeUsers", "newUsers"],
                ),
            )[0]
            period_events = retry_call(
                "GA4 period events report",
                lambda: ga_client.run_report(
                    start_date=period_start,
                    end_date=period_end,
                    dimensions=["eventName"],
                    metrics=["eventCount", "totalUsers"],
                    order_bys=[{"metric": {"metricName": "eventCount"}, "desc": True}],
                    limit=20,
                ),
            )
            period_sources = retry_call(
                "GA4 period sources report",
                lambda: ga_client.run_report(
                    start_date=period_start,
                    end_date=period_end,
                    dimensions=["sessionDefaultChannelGroup"],
                    metrics=["sessions"],
                    order_bys=[{"metric": {"metricName": "sessions"}, "desc": True}],
                    limit=10,
                ),
            )
            countries_raw = retry_call(
                "GA4 countries report",
                lambda: ga_client.run_report(
                    start_date=period_start,
                    end_date=period_end,
                    dimensions=["country"],
                    metrics=["activeUsers"],
                    order_bys=[{"metric": {"metricName": "activeUsers"}, "desc": True}],
                    limit=8,
                ),
            )
            page_engagement_raw = retry_call(
                "GA4 page engagement report",
                lambda: ga_client.run_report(
                    start_date=period_start,
                    end_date=period_end,
                    dimensions=["pageTitle"],
                    metrics=[
                        "screenPageViews",
                        "sessions",
                        "activeUsers",
                        "averageSessionDuration",
                        "bounceRate",
                    ],
                    order_bys=[{"metric": {"metricName": "screenPageViews"}, "desc": True}],
                    limit=30,
                ),
            )
            cta_breakdown_raw = retry_call(
                "GA4 CTA breakdown report",
                lambda: ga_client.run_report(
                    start_date=period_start,
                    end_date=period_end,
                    dimensions=["customEvent:event_label"],
                    metrics=["eventCount"],
                    dimension_filter={
                        "filter": {
                            "fieldName": "eventName",
                            "stringFilter": {"matchType": "EXACT", "value": "cta_click"},
                        }
                    },
                    order_bys=[{"metric": {"metricName": "eventCount"}, "desc": True}],
                    limit=12,
                ),
            )

            page_engagement = build_page_engagement(page_engagement_raw)
            page_lookup = {page["page"]: page for page in page_engagement}
            period_event_counts = {
                row["dimensions"][0]: to_int(row["metrics"][0]) for row in period_events
            }
            period_event_users = {
                row["dimensions"][0]: to_int(row["metrics"][1]) for row in period_events
            }
            ga4_period = {
                "start": period_start.isoformat(),
                "end": period_end.isoformat(),
                "totalSessions": to_int(period_overall["metrics"][0]),
                "totalPageViews": to_int(period_overall["metrics"][1]),
                "totalUsers": to_int(period_overall["metrics"][2]),
                "totalNewUsers": to_int(period_overall["metrics"][3]),
                "totalCTAClicks": period_event_counts.get("cta_click", 0),
                "totalLeads": period_event_counts.get("generate_lead", 0),
                "countries": [
                    {"country": row["dimensions"][0], "users": to_int(row["metrics"][0])}
                    for row in countries_raw
                ],
                "trafficSources": bucket_sources(
                    period_sources, to_int(period_overall["metrics"][0]), percent=False
                ),
                "funnel": [
                    {
                        "step": "Landed on site",
                        "users": to_int(period_overall["metrics"][2]),
                        "description": "Active users who visited the site",
                    },
                    {
                        "step": "Viewed homepage",
                        "users": page_lookup.get("Homepage", {}).get("activeUsers", 0),
                        "description": "Viewed the homepage",
                    },
                    {
                        "step": "Visited contact page",
                        "users": page_lookup.get("Contact", {}).get("activeUsers", 0),
                        "description": "Viewed the contact page",
                    },
                    {
                        "step": "Clicked a CTA",
                        "users": period_event_users.get("cta_click", 0),
                        "description": "Triggered a CTA click event",
                    },
                    {
                        "step": "Became a lead",
                        "users": period_event_users.get("generate_lead", 0),
                        "description": "Triggered a generate_lead event",
                    },
                ],
                "pageBounceRates": [
                    {"page": page["page"], "bounceRate": page["bounceRate"]}
                    for page in page_engagement
                ],
                "ctaBreakdown": [
                    {
                        "button": humanise_cta_label(row["dimensions"][0]),
                        "clicks": to_int(row["metrics"][0]),
                    }
                    for row in cta_breakdown_raw
                    if row["dimensions"][0] and row["dimensions"][0] != "(not set)"
                ],
                "navigationFlow": reports.get("ga4Period", {}).get("navigationFlow", {}),
                "pageEngagement": page_engagement,
                "exitPages": build_exit_pages(page_engagement),
                "seoInsights": reports.get("ga4Period", {}).get("seoInsights", {}),
            }
        except Exception as exc:
            print(
                "Warning: GA4 refresh failed. "
                f"Trying local CSV fallback. ({exc})"
            )
            ga4_period, weekly_ga = load_ga4_fallback(reports)

        if search_console_client is not None:
            try:
                seo_insights = retry_call(
                    "Search Console refresh",
                    lambda: refresh_search_console(
                        search_console_client, start_date=period_start, end_date=period_end
                    ),
                )
            except Exception as exc:
                print(
                    "Warning: Search Console refresh failed. "
                    f"Trying local CSV fallback. ({exc})"
                )
                try:
                    seo_insights = load_search_console_fallback(
                        start_date=period_start, end_date=period_end
                    )
                except Exception as fallback_exc:
                    print(
                        "Warning: Search Console fallback load failed. "
                        f"Keeping previous SEO insights. ({fallback_exc})"
                    )
                    seo_insights = reports.get("ga4Period", {}).get("seoInsights", {})
        else:
            try:
                seo_insights = load_search_console_fallback(
                    start_date=period_start, end_date=period_end
                )
            except Exception:
                seo_insights = reports.get("ga4Period", {}).get("seoInsights", {})

        refresh_date = date.today()
        reports["lastUpdated"] = refresh_date.isoformat()
        reports["dataThrough"] = latest_report_date.isoformat()
        ga4_period["seoInsights"] = seo_insights
        reports["ga4Period"] = ga4_period
        reports["linkedinYear"] = linkedin_year
        reports["topLinkedInPosts"] = top_linkedin_posts

        try:
            email_match = email_future.result()
        except Exception as exc:
            print(f"Warning: Email stats polling failed: {exc}")
            email_match = None

    while EMAIL_INTERACTIVE_WAIT and not email_match and sys.stdin.isatty():
        answer = input(
            "Email stats file not found yet. Keep waiting 10 more minutes? [y/N]: "
        ).strip().lower()
        if answer not in {"y", "yes"}:
            break
        email_match = poll_for_email_stats(time.time(), wait_minutes=10)

    if email_match:
        email_file, parsed_email = email_match
        email_stats = merge_email_metrics(email_defaults, parsed_email)
        print(f"Using email stats from: {email_file}")
    else:
        email_stats = email_defaults
        print(
            "No parseable email stats file found in Downloads. "
            "Keeping previous email values."
        )

    reports["weeks"] = [
        {
            "weekEnding": week_label,
            "ga": weekly_ga,
            "linkedin": linkedin_week["linkedin"],
            "email": email_stats,
        }
    ] + [week for week in reports.get("weeks", []) if week.get("weekEnding") != week_label]

    with REPORTS_PATH.open("w", encoding="utf-8") as handle:
        json.dump(reports, handle, indent=2, ensure_ascii=False)
        handle.write("\n")

    print("Updated reports.json")
    print(f"  Last updated: {reports['lastUpdated']}")
    print(
        "  Latest GA week:",
        weekly_ga["sessions"],
        "sessions,",
        weekly_ga["pageViews"],
        "page views,",
        weekly_ga["leads"],
        "leads",
    )
    print(
        "  Latest LinkedIn week:",
        linkedin_week["linkedin"]["followers"],
        "followers,",
        linkedin_week["linkedin"]["impressions"],
        "impressions,",
        f"{linkedin_week['linkedin']['engagementRate']}% engagement",
    )
    print(
        "  Email:",
        email_stats["campaignsSent"],
        "campaigns,",
        email_stats["totalRecipients"],
        "recipients,",
        f"{email_stats['openRate']}% open,",
        f"{email_stats['clickRate']}% click,",
        email_stats["unsubscribes"],
        "unsubscribes",
    )


if __name__ == "__main__":
    main()
