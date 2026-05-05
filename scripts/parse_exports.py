import openpyxl
from datetime import datetime, timedelta

wb = openpyxl.load_workbook('src/Exports/Content_2025-11-22_2026-02-19_RicWilson.xlsx')

ws = wb['ENGAGEMENT']
engagement = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    date_str, impressions, engagements = row
    if date_str and impressions is not None:
        d = datetime.strptime(date_str, '%m/%d/%Y')
        engagement[d] = {'impressions': impressions, 'engagements': engagements}

ws2 = wb['FOLLOWERS']
total_followers = 1336
followers = {}
for row in ws2.iter_rows(min_row=4, values_only=True):
    date_str, new_f = row
    if date_str and new_f is not None:
        d = datetime.strptime(date_str, '%m/%d/%Y')
        followers[d] = new_f

def get_week_ending(d):
    days_until_sunday = (6 - d.weekday()) % 7
    return d + timedelta(days=days_until_sunday)

weekly = {}
for d, data in engagement.items():
    we = get_week_ending(d)
    we_str = we.strftime('%Y-%m-%d')
    if we_str not in weekly:
        weekly[we_str] = {'impressions': 0, 'engagements': 0, 'newFollowers': 0}
    weekly[we_str]['impressions'] += data['impressions']
    weekly[we_str]['engagements'] += data['engagements']

for d, new_f in followers.items():
    we = get_week_ending(d)
    we_str = we.strftime('%Y-%m-%d')
    if we_str not in weekly:
        weekly[we_str] = {'impressions': 0, 'engagements': 0, 'newFollowers': 0}
    weekly[we_str]['newFollowers'] += new_f

sorted_weeks = sorted(weekly.keys(), reverse=True)
running_total = total_followers
for we_str in sorted_weeks:
    weekly[we_str]['followers'] = int(running_total)
    running_total -= weekly[we_str]['newFollowers']

for we_str in sorted_weeks:
    w = weekly[we_str]
    if w['impressions'] > 0:
        w['engagementRate'] = round(w['engagements'] / w['impressions'] * 100, 1)
    else:
        w['engagementRate'] = 0

print("LINKEDIN WEEKLY DATA:")
for we_str in sorted_weeks[:13]:
    w = weekly[we_str]
    print(f"  {we_str}: followers={w['followers']}, new={int(w['newFollowers'])}, impressions={int(w['impressions'])}, engagements={int(w['engagements'])}, rate={w['engagementRate']}%")
