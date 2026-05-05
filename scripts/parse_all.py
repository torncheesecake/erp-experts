import json
from datetime import datetime, timedelta
import openpyxl

# ─── GA4 SNAPSHOT-4: Daily data Jan 1 - Feb 18 2026 ───
lines = open("src/Exports/Reports_snapshot-4.csv").read().split("\n")

daily_users = {}
daily_new_users = {}
daily_engagement = {}

metric = None
year = None
i = 0
while i < len(lines):
    line = lines[i].strip()
    if line.startswith("# Start date:"):
        year_str = line.split(":")[1].strip()[:4]
        year = int(year_str)
    elif line.startswith("Nth day,"):
        metric = line.split(",", 1)[1].strip()
    elif line and not line.startswith("#") and "," in line and metric and year == 2026:
        parts = line.split(",")
        if parts[0].isdigit():
            day_num = int(parts[0])
            val = float(parts[1]) if parts[1] else 0
            date = datetime(2026, 1, 1) + timedelta(days=day_num)
            ds = date.strftime("%Y-%m-%d")
            if metric == "Active users":
                daily_users[ds] = val
            elif metric == "New users":
                daily_new_users[ds] = val
            elif "engagement time" in metric.lower():
                daily_engagement[ds] = val
    elif line.startswith("# Where do") or line.startswith("# How are active"):
        break
    i += 1

# Bucket into weeks (Mon-Sun)
def get_week_ending(date_str):
    d = datetime.strptime(date_str, "%Y-%m-%d")
    days_until_sunday = (6 - d.weekday()) % 7
    if days_until_sunday == 0 and d.weekday() == 6:
        return d.strftime("%Y-%m-%d")
    return (d + timedelta(days=days_until_sunday)).strftime("%Y-%m-%d")

weeks = {}
for ds in sorted(daily_users.keys()):
    we = get_week_ending(ds)
    if we not in weeks:
        weeks[we] = {"users": 0, "new_users": 0, "engagement_sum": 0, "engagement_count": 0, "days": 0}
    weeks[we]["users"] += int(daily_users.get(ds, 0))
    weeks[we]["new_users"] += int(daily_new_users.get(ds, 0))
    eng = daily_engagement.get(ds, 0)
    if eng > 0:
        weeks[we]["engagement_sum"] += eng
        weeks[we]["engagement_count"] += 1
    weeks[we]["days"] += 1

print("=== GA4 Weekly Buckets ===")
for we in sorted(weeks.keys()):
    w = weeks[we]
    avg_eng = w["engagement_sum"] / w["engagement_count"] if w["engagement_count"] > 0 else 0
    print(f"{we}: users={w['users']}, new={w['new_users']}, avg_eng={avg_eng:.0f}s, days={w['days']}")

# Parse traffic sources
print("\n=== Traffic Sources 2026 ===")
in_section = False
is_2026 = False
for line in lines:
    line = line.strip()
    if "Session primary channel group" in line:
        in_section = True
        continue
    if in_section and line.startswith("# Start date: 2026"):
        is_2026 = True
    if in_section and line.startswith("# Start date: 2025"):
        is_2026 = False
        in_section = False
    if in_section and is_2026 and line and not line.startswith("#"):
        parts = line.split(",")
        if len(parts) == 2 and parts[1].strip().isdigit():
            print(f"  {parts[0]}: {parts[1]}")

# Parse country
print("\n=== Countries 2026 ===")
found = False
in_section = False
for line in lines:
    line = line.strip()
    if line == "Country,Active users" and not found:
        found = True
        in_section = True
        continue
    if in_section:
        if line.startswith("#") or line == "":
            in_section = False
            continue
        parts = line.split(",")
        if len(parts) == 2 and parts[1].strip().isdigit() and int(parts[1]) > 0:
            print(f"  {parts[0]}: {parts[1]}")

# Events
print("\n=== Events 2026 ===")
found_events = False
in_section = False
for line in lines:
    line = line.strip()
    if line == "Event name,Event count" and not found_events:
        found_events = True
        in_section = True
        continue
    if in_section:
        if line.startswith("#") or line == "":
            break
        parts = line.split(",")
        if len(parts) == 2:
            print(f"  {parts[0]}: {parts[1]}")

# ─── LinkedIn ───
print("\n=== LinkedIn Demographics ===")
wb = openpyxl.load_workbook("src/Exports/Content_2025-02-20_2026-02-19_RicWilson.xlsx")
ws = wb["DEMOGRAPHICS"]
rows = list(ws.iter_rows(values_only=True))
for r in rows[1:]:
    print(f"  {r[0]}: {r[1]} ({r[2]*100:.1f}%)")

ws = wb["DISCOVERY"]
rows = list(ws.iter_rows(values_only=True))
print("\n=== LinkedIn Discovery ===")
for r in rows[1:]:
    print(f"  {r[0]}: {r[1]}")

# Top posts
ws = wb["TOP POSTS"]
rows = list(ws.iter_rows(values_only=True))
print("\n=== LinkedIn Top Posts (by engagement) ===")
for r in rows[3:13]:
    url = r[0] if r[0] else ""
    date = r[1] if r[1] else ""
    eng = r[2] if r[2] else 0
    imp_url = r[4] if r[4] else ""
    imp_date = r[5] if r[5] else ""
    imp = r[6] if r[6] else 0
    print(f"  Date: {date}, Engagements: {eng}, Impressions: {imp}")

# Daily engagement for recent weeks
ws = wb["ENGAGEMENT"]
rows = list(ws.iter_rows(values_only=True))
print("\n=== LinkedIn Daily Engagement (last 30 days) ===")
for r in rows[-30:]:
    print(f"  {r[0]}: imp={r[1]}, eng={r[2]}")

# Follower data
ws = wb["FOLLOWERS"]
rows = list(ws.iter_rows(values_only=True))
print(f"\n=== LinkedIn Followers ===")
print(f"  Total: {rows[0][1]}")
print("  Last 14 days:")
for r in rows[-14:]:
    print(f"    {r[0]}: +{r[1]}")
